"""
====================================================
Project Name: Google Suggestion Excel Automation
Author: Sohel Tanvir
Description:
    This Python script automates the process of:
    - Reading keywords from an Excel file
    - Fetching Google search suggestions using Selenium
    - Finding the longest and shortest suggestions
    - Writing the results back into the Excel file

Technologies Used:
    - Python
    - Selenium WebDriver
    - OpenPyXL

File Requirements:
    - excel_file.xlsx (must exist)
    - Sheet name should match the current weekday (e.g., Monday, Tuesday)

====================================================
"""

import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def setup_browser():
    """
    Initializes the Chrome WebDriver with required options.

    Returns:
        driver (webdriver): Selenium Chrome WebDriver instance
    """
    chrome_options = Options()
    chrome_options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--lang=en-US")
    chrome_options.add_argument("--disable-gpu")

    service = Service(r'C:\Users\Dream\PycharmProjects\pythonProject\Demo\chromedriver.exe')
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver


def get_today_sheet(workbook):
    """
    Gets the worksheet based on the current day name.

    Args:
        workbook (Workbook): Loaded Excel workbook

    Returns:
        sheet (Worksheet): Matching worksheet object
    """
    today = datetime.datetime.now()
    day_name = today.strftime("%A")

    print("Today is:", day_name)

    if day_name not in workbook.sheetnames:
        raise Exception(f"Sheet '{day_name}' not found in Excel file.")

    return workbook[day_name]


def get_google_suggestions(driver, keyword):
    """
    Fetches Google auto-suggestions for a given keyword.

    Args:
        driver (webdriver): Selenium WebDriver
        keyword (str): Search keyword

    Returns:
        list: A list of suggestion strings
    """
    try:
        driver.get("https://www.google.com/?hl=en")

        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q"))
        )
        search_box.clear()
        search_box.send_keys(keyword)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li.sbct"))
        )

        elements = driver.find_elements(By.CSS_SELECTOR, "li.sbct span")

        return [el.text.strip() for el in elements if el.text.strip()]

    except Exception as e:
        print(f"Error while fetching suggestions for '{keyword}':", e)
        return []


def find_longest_shortest(suggestions):
    """
    Finds the longest and shortest text from a list.

    Args:
        suggestions (list): List of suggestion strings

    Returns:
        tuple: (longest_string, shortest_string)
    """
    if not suggestions:
        return None, None

    longest = max(suggestions, key=len)
    shortest = min(suggestions, key=len)
    return longest, shortest


def main():
    """
    Main function that controls the workflow:
    - Load Excel file
    - Read keywords row by row
    - Fetch Google suggestions
    - Write longest and shortest suggestions back to Excel
    """
    try:
        driver = setup_browser()
        workbook = load_workbook("excel_file.xlsx")
        sheet = get_today_sheet(workbook)

        for row in range(2, sheet.max_row + 1):
            keyword = sheet.cell(row=row, column=1).value

            if not keyword:
                continue

            print(f"\nProcessing keyword: {keyword}")

            suggestions = get_google_suggestions(driver, keyword)
            longest, shortest = find_longest_shortest(suggestions)

            print("Longest Suggestion:", longest or "None")
            print("Shortest Suggestion:", shortest or "None")

            sheet.cell(row=row, column=2).value = longest
            sheet.cell(row=row, column=3).value = shortest

        workbook.save("excel_file.xlsx")

    except Exception as error:
        print("Script failed:", error)

    finally:
        driver.quit()
        print("\nBrowser closed. Script completed successfully.")


# ===== Entry Point =====
if __name__ == "__main__":
    main()
